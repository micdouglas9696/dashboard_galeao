# -*- coding: utf-8 -*-
"""
SESCINC SBGL Dashboard — Complete Seed Data Generator
Parses ALL Excel spreadsheets (Janeiro-Junho) and generates updated seed-data.js
with proper 'mes' field on ALL records to avoid duplication.
"""
import openpyxl
import datetime
import json
import re
import os
import unicodedata

# ────────── Helpers ──────────

def normalize_funcao(fun):
    if not fun:
        return 'BA'
    fun = str(fun).strip().upper()
    # Remove non-breaking spaces
    fun = fun.replace('\xa0', ' ').strip()
    
    if fun in ['BA', 'B.A.', 'ba', 'B.A']:
        return 'BA'
    if fun in ['BA2', 'BA 2', 'BA-2', 'BA-02', 'B.A II']:
        return 'BA2'
    if fun in ['BA-MC', 'BAMC', 'BA MC', 'BA- MC']:
        return 'BA-MC'
    if fun in ['BA-LR', 'BA LR']:
        return 'BA-LR'
    if fun in ['BA-RE', 'BA RE']:
        return 'BA-RE'
    if fun in ['BA-CE', 'BA CE']:
        return 'BA-CE'
    if fun in ['BA-MA', 'BA MA']:
        return 'BA-MA'
    if fun in ['OC']:
        return 'OC'
    
    if re.search(r'BA.*MC', fun): return 'BA-MC'
    if re.search(r'BA.*LR', fun): return 'BA-LR'
    if re.search(r'BA.*RE', fun): return 'BA-RE'
    if re.search(r'BA.*CE', fun): return 'BA-CE'
    if re.search(r'BA.*MA', fun): return 'BA-MA'
    if '2' in fun or 'II' in fun: return 'BA2'
    return 'BA'


def normalize_name(name):
    if not name:
        return ''
    name = str(name).strip().upper()
    name = name.replace('\xa0', ' ')
    name = re.sub(r'\s+', ' ', name)
    name = re.sub(r'\.$', '', name)
    name = ''.join(c for c in unicodedata.normalize('NFD', name) if unicodedata.category(c) != 'Mn')
    return name


def clean_numeric(val):
    """Clean numeric values that might have non-breaking spaces."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val) if isinstance(val, float) and val == int(val) else val
    s = str(val).replace('\xa0', '').strip()
    if not s or s.upper() == 'NR':
        return None
    try:
        return int(float(s))
    except:
        return None


def clean_corrida(val):
    """Clean and normalize corrida (run time) field."""
    if val is None:
        return '', None
    
    s = str(val).replace('\xa0', '').strip()
    if not s:
        return '', None
    if s.upper() in ['NR', 'FERIAS', 'FÉRIAS']:
        return s.upper(), None
    
    # Handle special quote characters
    s = s.replace('\u2018', "'").replace('\u2019', "'").replace('\u201c', '"').replace('\u201d', '"')
    s = s.replace('\u2032', "'").replace('\u2033', '"')
    
    # Format: distance like "2.450m" or "2.460m" or "2,450m"
    m_dist = re.match(r'^(\d+)[.,](\d+)\s*m$', s, re.IGNORECASE)
    if m_dist:
        # This is a distance in meters, keep as-is (no seconds conversion)
        return s, None
    
    # Format: mm'ss" or mm'ss'' (various quote styles)
    m = re.match(r"(\d+)['\u2019](\d+)[\"'\u2019\u201d]*\s*$", s)
    if m:
        mins = int(m.group(1))
        secs = int(m.group(2))
        return f"{mins}'{secs:02d}\"", mins * 60 + secs
    
    # Format: HH:MM:SS (e.g., 00:11:49)
    parts = s.split(':')
    if len(parts) == 3:
        try:
            h, mi, sec = int(parts[0]), int(parts[1]), int(parts[2])
            total = h * 3600 + mi * 60 + sec
            actual_mins = mi if h == 0 else h * 60 + mi
            return f"{actual_mins}'{sec:02d}\"", total
        except:
            pass
    elif len(parts) == 2:
        try:
            mi, sec = int(parts[0]), int(parts[1])
            return f"{mi}'{sec:02d}\"", mi * 60 + sec
        except:
            pass
    
    return s, None


MONTH_MAP = {
    'JANEIRO': 'Janeiro', 'FEVEREIRO': 'Fevereiro', 'MARÇO': 'Março',
    'MARCO': 'Março', 'ABRIL': 'Abril', 'MAIO': 'Maio', 'JUNHO': 'Junho',
    'JULHO': 'Julho', 'AGOSTO': 'Agosto', 'SETEMBRO': 'Setembro',
    'OUTUBRO': 'Outubro', 'NOVEMBRO': 'Novembro', 'DEZEMBRO': 'Dezembro'
}

MONTH_INDEX = {
    'Janeiro': 0, 'Fevereiro': 1, 'Março': 2, 'Abril': 3, 'Maio': 4, 'Junho': 5,
    'Julho': 6, 'Agosto': 7, 'Setembro': 8, 'Outubro': 9, 'Novembro': 10, 'Dezembro': 11
}


def detect_month_from_filename(filename):
    """Extract month name from filename."""
    # macOS returns filenames in NFD; normalize to NFC for comparison
    fname_upper = unicodedata.normalize('NFC', filename).upper()
    for key, val in MONTH_MAP.items():
        if key in fname_upper:
            return val
    return None


def detect_month_from_sheet(sheet_name):
    """Extract month name from sheet name."""
    sn_upper = unicodedata.normalize('NFC', sheet_name).strip().upper()
    for key, val in MONTH_MAP.items():
        if key == sn_upper or key in sn_upper:
            return val
    return None


# ────────── TAF Parser ──────────

def parse_taf_file(filepath, month_name):
    """Parse a single TAF XLSX file and return records with 'mes' field."""
    print(f'  Parsing TAF: {os.path.basename(filepath)} -> {month_name}')
    wb = openpyxl.load_workbook(filepath, data_only=True)
    records = []
    
    # Try to find the right sheet
    ws = None
    for sn in wb.sheetnames:
        if 'gráfico' in sn.lower() or 'grafico' in sn.lower():
            continue
        ws = wb[sn]
        break
    
    if ws is None:
        wb.close()
        return records
    
    # Row 5 headers: NOME, EQUIPE, FUNÇÃO, IDADE, Flexão Solo, Abdominal, Barra Fixa, Corrida, Resultado
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        nome = row[0]
        if not nome or str(nome).strip() == '':
            continue
        nome_str = str(nome).strip()
        
        # Skip summary/total rows
        nome_upper = nome_str.upper()
        if any(kw in nome_upper for kw in ['TOTAL', 'CONTAGEM', 'SATISFAT', 'INSATISF']):
            continue
        
        equipe = str(row[1]).strip().upper() if row[1] else ''
        if not equipe or equipe == 'NONE':
            continue
            
        funcao = normalize_funcao(row[2])
        
        # Handle idade - might have non-breaking spaces
        idade = clean_numeric(row[3])
        
        # Handle FÉRIAS status
        is_ferias = False
        for cell_val in row[4:8]:
            if cell_val and 'FERIAS' in str(cell_val).upper().replace('É', 'E'):
                is_ferias = True
        
        if is_ferias:
            status = 'ferias'
            flexao = None
            abdominal = None
            barra = None
            corrida = 'FÉRIAS'
            corrida_seconds = None
        else:
            flexao = clean_numeric(row[4])
            abdominal = clean_numeric(row[5])
            barra = clean_numeric(row[6])
            
            corrida_raw = row[7]
            corrida, corrida_seconds = clean_corrida(corrida_raw)
            
            status = 'ok'
            if corrida.upper() == 'NR' or (flexao is None and abdominal is None and barra is None and not corrida):
                status = 'nr'
        
        resultado = str(row[8]).strip() if row[8] else 'Satisfatório'
        # Normalize resultado
        if 'insatisf' in resultado.lower():
            resultado = 'Insatisfatório'
        elif 'satisf' in resultado.lower():
            resultado = 'Satisfatório'
        
        records.append({
            'nome': nome_str,
            'equipe': equipe,
            'funcao': funcao,
            'idade': idade,
            'flexao': flexao,
            'abdominal': abdominal,
            'barra': barra,
            'corrida': corrida,
            'corridaSeconds': corrida_seconds,
            'resultado': resultado,
            'status': status,
            'mes': month_name
        })
    
    wb.close()
    return records


# ────────── TP-EPR Parser ──────────

def parse_tpepr_file(filepath, month_name):
    """Parse a single TP-EPR XLSX file and return records with 'mes' field."""
    print(f'  Parsing TP-EPR: {os.path.basename(filepath)} -> {month_name}')
    wb = openpyxl.load_workbook(filepath, data_only=True)
    records = []
    
    # Try to find the right sheet
    ws = None
    for sn in wb.sheetnames:
        if 'gráfico' in sn.lower() or 'grafico' in sn.lower():
            continue
        ws = wb[sn]
        break
    
    if ws is None:
        wb.close()
        return records
    
    # Auto-detect column offset by checking header row (row 5)
    # Some files have 6 cols (empty first col, data starts at B) 
    # Some files have 5 cols (data starts at A)
    header_row = [ws.cell(row=5, column=c).value for c in range(1, 8)]
    
    # Detect offset: if col A is None or col A contains 'NOME'
    offset = 0  # default: no offset (5 cols)
    if header_row[0] is None and header_row[1] and 'NOME' in str(header_row[1]).upper():
        offset = 1  # 6-col layout with empty first column
    elif header_row[0] and 'NOME' in str(header_row[0]).upper():
        offset = 0  # 5-col layout
    else:
        # Fallback: check if first cell of row 6 is None
        first_data = ws.cell(row=6, column=1).value
        if first_data is None:
            offset = 1
    
    print(f'    Column offset: {offset} ({"6-col" if offset else "5-col"} layout)')
    
    # Column indices based on offset
    col_nome = offset  # 0 or 1
    col_equipe = offset + 1
    col_funcao = offset + 2
    col_tempo = offset + 3
    col_resultado = offset + 4
    
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        # Ensure row has enough columns
        row = list(row) + [None] * max(0, col_resultado + 1 - len(row))
        
        nome = row[col_nome]
        if not nome or str(nome).strip() == '':
            continue
        nome_str = str(nome).strip()
        
        # Skip summary rows
        nome_upper = nome_str.upper()
        if any(kw in nome_upper for kw in ['TOTAL', 'CONTAGEM', 'EXCELENTE', 'BOM', 'INSATISF']):
            continue
        
        equipe = str(row[col_equipe]).strip().upper() if row[col_equipe] else ''
        if not equipe or equipe == 'NONE':
            continue
            
        funcao = normalize_funcao(row[col_funcao])
        tempo_raw = row[col_tempo] if len(row) > col_tempo else None
        resultado = str(row[col_resultado]).strip() if len(row) > col_resultado and row[col_resultado] else ''
        
        tempo_seconds = 0
        tempo_formatted = ''
        
        if tempo_raw:
            tempo_str = str(tempo_raw).strip().upper().replace('\xa0', '')
            
            if 'FERIAS' in tempo_str or 'FÉRIAS' in tempo_str:
                tempo_formatted = 'FÉRIAS'
                tempo_seconds = 0
            elif isinstance(tempo_raw, datetime.time):
                tempo_seconds = tempo_raw.hour * 3600 + tempo_raw.minute * 60 + tempo_raw.second
                tempo_formatted = f"{tempo_raw.minute:02d}:{tempo_raw.second:02d}"
            elif isinstance(tempo_raw, (int, float)):
                total_sec = int(round(tempo_raw * 86400))
                tempo_seconds = total_sec
                minutos = total_sec // 60
                segundos = total_sec % 60
                tempo_formatted = f"{minutos:02d}:{segundos:02d}"
            else:
                parts = tempo_str.split(':')
                if len(parts) == 3:
                    try:
                        tempo_seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                        tempo_formatted = f"{int(parts[1]):02d}:{int(parts[2]):02d}"
                    except:
                        tempo_formatted = str(tempo_raw).strip()
                elif len(parts) == 2:
                    try:
                        tempo_seconds = int(parts[0]) * 60 + int(parts[1])
                        tempo_formatted = f"{int(parts[0]):02d}:{int(parts[1]):02d}"
                    except:
                        tempo_formatted = str(tempo_raw).strip()
                else:
                    tempo_formatted = str(tempo_raw).strip()
        
        # Determine resultado if missing
        if not resultado or resultado == 'None':
            if tempo_formatted == 'FÉRIAS':
                resultado = 'Férias'
            elif tempo_seconds <= 60:
                resultado = 'Excelente'
            elif tempo_seconds <= 90:
                resultado = 'Bom'
            else:
                resultado = 'Insatisfatório'
        
        # Normalize resultado
        if 'ruim' in resultado.lower() or 'insatisf' in resultado.lower():
            resultado = 'Insatisfatório'
        
        records.append({
            'nome': nome_str,
            'equipe': equipe,
            'funcao': funcao,
            'tempoSeconds': tempo_seconds,
            'tempoFormatted': tempo_formatted,
            'resultado': resultado,
            'mes': month_name
        })
    
    wb.close()
    return records


# ────────── TR Parser (unchanged from original) ──────────

def parse_tr():
    f = os.path.join(BASE_DIR, 'DESEMPENHO DA EXECUÇÃO TR.xlsx')
    print(f'Parsing TR: {os.path.basename(f)}')
    wb = openpyxl.load_workbook(f, data_only=True)
    records = []
    
    month_indices = {
        'Janeiro': 0, 'Fevereiro': 1, 'Março': 2, 'Abril': 3, 'Maio': 4, 'Junho': 5
    }
    
    for sn in ['CABECEIRA 28', 'CABECEIRA 33', 'CABECEIRA 15']:
        ws = wb[sn]
        cabeceira = sn.split()[-1]
        
        col_mappings = []
        current_month = None
        
        row7 = [ws.cell(row=7, column=c).value for c in range(1, ws.max_column+1)]
        row8 = [ws.cell(row=8, column=c).value for c in range(1, ws.max_column+1)]
        
        for col_idx in range(1, len(row7)):
            m_val = row7[col_idx]
            if m_val and str(m_val).strip().upper() in MONTH_MAP:
                current_month = MONTH_MAP[str(m_val).strip().upper()]
            
            cci_val = row8[col_idx]
            if cci_val and current_month:
                cci_str = str(cci_val).strip()
                if 'CCI' in cci_str or any(char.isdigit() for char in cci_str):
                    col_mappings.append({
                        'col': col_idx + 1,
                        'month': current_month,
                        'cci': cci_str
                    })
        
        for r_idx in range(9, 13):
            equipe = str(ws.cell(row=r_idx, column=1).value).strip().upper()
            if not equipe or equipe == 'NONE':
                continue
                
            for mapping in col_mappings:
                c_idx = mapping['col']
                val = ws.cell(row=r_idx, column=c_idx).value
                
                status = 'ok'
                tempo_seconds = None
                tempo_formatted = None
                
                val_str = str(val).strip().upper() if val is not None else ''
                
                if val_str == 'NR':
                    status = 'nr'
                    tempo_formatted = 'NR'
                elif val_str == 'X' or val_str == '':
                    status = 'na'
                else:
                    status = 'ok'
                    if isinstance(val, datetime.time):
                        tempo_seconds = val.hour * 60 + val.minute
                        tempo_formatted = f"{val.hour:02d}:{val.minute:02d}"
                    elif isinstance(val, (int, float)):
                        total_sec = int(round(val * 1440))
                        tempo_seconds = total_sec
                        tempo_formatted = f"{total_sec // 60:02d}:{total_sec % 60:02d}"
                    else:
                        parts = val_str.split(':')
                        if len(parts) >= 2:
                            try:
                                if len(parts) == 3:
                                    tempo_seconds = int(parts[0]) * 60 + int(parts[1])
                                    tempo_formatted = f"{int(parts[0]):02d}:{int(parts[1]):02d}"
                                else:
                                    tempo_seconds = int(parts[0]) * 60 + int(parts[1])
                                    tempo_formatted = f"{int(parts[0]):02d}:{int(parts[1]):02d}"
                            except:
                                status = 'empty'
                        else:
                            status = 'empty'
                
                if status != 'na' and status != 'empty':
                    records.append({
                        'cabeceira': cabeceira,
                        'equipe': equipe,
                        'mes': mapping['month'],
                        'mesIndex': month_indices.get(mapping['month'], 0),
                        'cci': mapping['cci'],
                        'tempoFormatted': tempo_formatted,
                        'tempoSeconds': tempo_seconds,
                        'status': status
                    })
    wb.close()
    return records


# ────────── Teórica Parser (unchanged from original) ──────────

def parse_teorica(colaborador_map):
    f = os.path.join(BASE_DIR, '_Aplicação de Avaliação Teórica PTR-BA 2º Trimestre - 2026. (1-87).xlsx')
    print(f'Parsing Teórica: {os.path.basename(f)}')
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb['Sheet1']
    records = []
    
    import difflib
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
            
        nome_original = str(row[8]).strip() if row[8] else ''
        nome_norm = normalize_name(nome_original)
        
        funcao_orig = str(row[11]).strip() if row[11] else ''
        funcao = normalize_funcao(funcao_orig)
        
        nota = float(row[5]) if row[5] is not None else 0.0
        
        equipe = 'Não identificada'
        if nome_norm in colaborador_map:
            equipe = colaborador_map[nome_norm]['equipe']
        else:
            matches = difflib.get_close_matches(nome_norm, colaborador_map.keys(), n=1, cutoff=0.75)
            if matches:
                equipe = colaborador_map[matches[0]]['equipe']
                
        questoes = []
        q_idx = 1
        for col_idx in range(17, len(row), 3):
            if col_idx + 1 < len(row):
                pontos = row[col_idx + 1]
                if pontos is not None:
                    try:
                        questoes.append({
                            'num': q_idx,
                            'pontos': float(pontos)
                        })
                    except:
                        pass
                q_idx += 1
                
        records.append({
            'id': int(row[0]),
            'nome': nome_original,
            'funcao': funcao,
            'funcaoOriginal': funcao_orig,
            'aeroporto': 'SBGL',
            'nota': nota,
            'equipe': equipe,
            'questoes': questoes
        })
        
    wb.close()
    return records


# ────────── Main ──────────

BASE_DIR = '/Users/m.dbranding/Desktop/OSeas'
PLANILHAS_DIR = os.path.join(BASE_DIR, 'js', 'planilhas')


def main():
    all_taf_records = []
    all_tpepr_records = []
    
    # ── 1. Parse TAF files from planilhas folder (Janeiro-Maio) ──
    print('\n=== Parsing TAF files (planilhas folder) ===')
    taf_files = sorted([f for f in os.listdir(PLANILHAS_DIR) if 'TAF' in f.upper() and f.endswith('.xlsx') and not f.startswith('~')])
    for taf_file in taf_files:
        month = detect_month_from_filename(taf_file)
        if month:
            filepath = os.path.join(PLANILHAS_DIR, taf_file)
            records = parse_taf_file(filepath, month)
            all_taf_records.extend(records)
            print(f'    -> {len(records)} records for {month}')
    
    # ── 2. Parse TAF file from root (Junho) ──
    print('\n=== Parsing TAF Junho (root) ===')
    junho_taf = os.path.join(BASE_DIR, 'AFERIÇÃO - TAF JUNHO 2026.xlsx')
    if os.path.exists(junho_taf):
        records = parse_taf_file(junho_taf, 'Junho')
        all_taf_records.extend(records)
        print(f'    -> {len(records)} records for Junho')
    
    # ── 3. Parse TP-EPR files from planilhas folder (Janeiro-Maio) ──
    print('\n=== Parsing TP-EPR files (planilhas folder) ===')
    tpepr_files = sorted([f for f in os.listdir(PLANILHAS_DIR) if 'TP-EPR' in f.upper() and f.endswith('.xlsx') and not f.startswith('~')])
    for tpepr_file in tpepr_files:
        month = detect_month_from_filename(tpepr_file)
        if month:
            filepath = os.path.join(PLANILHAS_DIR, tpepr_file)
            records = parse_tpepr_file(filepath, month)
            all_tpepr_records.extend(records)
            print(f'    -> {len(records)} records for {month}')
    
    # ── 4. Parse TP-EPR file from root (Junho) ──
    print('\n=== Parsing TP-EPR Junho (root) ===')
    junho_tpepr = os.path.join(BASE_DIR, 'AFERIÇÃO TP-EPR JUNHO 2026.xlsx')
    if os.path.exists(junho_tpepr):
        records = parse_tpepr_file(junho_tpepr, 'Junho')
        all_tpepr_records.extend(records)
        print(f'    -> {len(records)} records for Junho')
    
    # ── 5. Deduplication check ──
    print('\n=== Deduplication Check ===')
    
    # TAF: deduplicate by (nome_normalized, mes)
    taf_seen = set()
    taf_deduped = []
    for r in all_taf_records:
        key = (normalize_name(r['nome']), r['mes'])
        if key not in taf_seen:
            taf_seen.add(key)
            taf_deduped.append(r)
        else:
            print(f'  TAF DUPLICATE SKIPPED: {r["nome"]} - {r["mes"]}')
    all_taf_records = taf_deduped
    
    # TPEPR: deduplicate by (nome_normalized, mes)
    tpepr_seen = set()
    tpepr_deduped = []
    for r in all_tpepr_records:
        key = (normalize_name(r['nome']), r['mes'])
        if key not in tpepr_seen:
            tpepr_seen.add(key)
            tpepr_deduped.append(r)
        else:
            print(f'  TPEPR DUPLICATE SKIPPED: {r["nome"]} - {r["mes"]}')
    all_tpepr_records = tpepr_deduped
    
    # ── 6. Parse TR ──
    print('\n=== Parsing TR ===')
    tr_records = parse_tr()
    print(f'    -> {len(tr_records)} TR records')
    
    # ── 7. Build colaborador map from ALL months for Teórica matching ──
    colaborador_map = {}
    for r in all_taf_records:
        n = normalize_name(r['nome'])
        if n and r['equipe']:
            colaborador_map[n] = {'equipe': r['equipe'], 'funcao': r['funcao']}
    for r in all_tpepr_records:
        n = normalize_name(r['nome'])
        if n and r['equipe'] and n not in colaborador_map:
            colaborador_map[n] = {'equipe': r['equipe'], 'funcao': r['funcao']}
    
    # ── 8. Parse Teórica ──
    print('\n=== Parsing Teórica ===')
    teorica_records = parse_teorica(colaborador_map)
    print(f'    -> {len(teorica_records)} Teórica records')
    
    # ── 9. Summary by month ──
    print('\n=== SUMMARY ===')
    taf_by_month = {}
    for r in all_taf_records:
        m = r['mes']
        taf_by_month[m] = taf_by_month.get(m, 0) + 1
    print(f'TAF records by month: {json.dumps(taf_by_month, ensure_ascii=False)}')
    
    tpepr_by_month = {}
    for r in all_tpepr_records:
        m = r['mes']
        tpepr_by_month[m] = tpepr_by_month.get(m, 0) + 1
    print(f'TPEPR records by month: {json.dumps(tpepr_by_month, ensure_ascii=False)}')
    
    print(f'TR records: {len(tr_records)}')
    print(f'Teórica records: {len(teorica_records)}')
    print(f'Total TAF: {len(all_taf_records)}')
    print(f'Total TPEPR: {len(all_tpepr_records)}')
    
    # ── 10. Generate seed-data.js ──
    seed_data = {
        'taf': {
            'records': all_taf_records,
            'uploadedAt': datetime.datetime.utcnow().isoformat() + 'Z'
        },
        'tpepr': {
            'records': all_tpepr_records,
            'uploadedAt': datetime.datetime.utcnow().isoformat() + 'Z'
        },
        'tr': {
            'records': tr_records,
            'uploadedAt': datetime.datetime.utcnow().isoformat() + 'Z'
        },
        'teorica': {
            'records': teorica_records,
            'uploadedAt': datetime.datetime.utcnow().isoformat() + 'Z'
        }
    }
    
    out_file = os.path.join(BASE_DIR, 'js', 'seed-data.js')
    
    with open(out_file, 'w', encoding='utf-8') as f:
        f.write("/**\n * SESCINC SBGL Dashboard — Seed Data\n * Automatically generated from Excel spreadsheets\n * Generated at: " + datetime.datetime.now().isoformat() + "\n */\n\n")
        f.write("window.SESCINC = window.SESCINC || {};\n")
        f.write("window.SESCINC.SeedData = ")
        json.dump(seed_data, f, ensure_ascii=False, indent=2)
        f.write(";\n")
    
    print(f'\n✅ Seed data written to {out_file}')
    print(f'   File size: {os.path.getsize(out_file):,} bytes')


if __name__ == '__main__':
    main()
